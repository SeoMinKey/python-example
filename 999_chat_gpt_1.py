import openpyxl
from openpyxl.utils import get_column_letter

# 새 엑셀 파일 생성
wb = openpyxl.Workbook()

# 첫 번째 시트 선택
ws = wb.active

# 첫 번째 행에 열 이름 넣기
header = ['이름', '성별', '나이', '개발여부']
for col_num, column_title in enumerate(header, 1):
    col_letter = get_column_letter(col_num)
    ws[f"{col_letter}1"] = column_title

# 데이터 입력
data = [
    ['김철수', '남', 28, 'O'],
    ['이영희', '여', 25, 'O'],
    ['박민수', '남', 31, 'X'],
    ['최지수', '여', 27, 'O'],
    ['홍길동', '남', 24, 'X']
]

# 데이터를 행 단위로 엑셀에 입력
for row_num, row_data in enumerate(data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}{row_num}"] = cell_value

# 엑셀 파일 저장
wb.save("수업인원.xlsx")
